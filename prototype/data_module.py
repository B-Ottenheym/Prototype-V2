import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from dictionaries import variable_labels

def generate_synthetic_data(n_samples):

    np.random.seed(31)
    #n_samples = 2000

    numerical_vars = {
        #Owner
        "payment_reliability": {"mean": 3.0, "std": 1.0, "min": 1, "max": 5},
        "decision_making_structure": {"mean": 4, "std": 2, "min": 1, "max": 10},
        "change_order_freq": {"mean": 7, "std": 5, "min": 0, "max": 30},
        #Contractor
        "contractor_experience": {"mean": 12, "std": 6, "min": 1, "max": 40},
        "num_subcontractors": {"mean": 8, "std": 6, "min": 1, "max": 40},
        "planning_detail": {"mean": 80, "std": 10, "min": 50, "max": 100},
        #Design
        "design_maturity": {"mean": 70, "std": 15, "min": 30, "max": 100},
        "pre_design_data_completeness": {"mean": 75, "std": 15, "min": 30, "max": 100},
        "design_team_experience_years": {"mean": 12, "std": 6, "min": 1, "max": 40},
        #Labour
        "labour_shortage": {"mean": 10, "std": 8, "min": 0, "max": 40},
        "experienced_labour": {"mean": 50, "std": 20, "min": 20, "max": 100},
        #Consultant
        "consultant_experience": {"mean": 10, "std": 5, "min": 1, "max": 30},
        "consultant_availability": {"mean": 80, "std": 15, "min": 30, "max": 100},
        #Materials
        "material_supply_reliability": {"mean": 3.5, "std": 1.0, "min": 1, "max": 5},
        "expected_inflation": {"mean": 4, "std": 2, "min": 0, "max": 10},
        #External
        "weather_risk": {"mean": 8, "std": 6, "min": 0, "max": 40},
        "subsurface_risk": {"mean": 3.0, "std": 1.0, "min": 1, "max": 5},
        "permitting_complexity": {"mean": 3.0, "std": 1.0, "min": 1, "max": 5},
        #Equipment
        "equipment_availability": {"mean": 0.85, "std": 0.15, "min": 0.3, "max": 1.0},
        "average_equipment_age": {"mean": 6, "std": 4, "min": 0, "max": 25}
    }

    categorical_vars = { #Also includes binary
        #Project
        "contract_award_method": (["Laagste prijs", "Laagste kosten op basis van kosteneffectiviteit", "Beste Prijs-Kwaliteitverhouding"], [0.5, 0.3, 0.2]),
        "contract_type": (["Lump-sum", "UAV-GC", "Kost+"], [0.6, 0.3, 0.1]),
        #Design
        "use_of_bim": ([0, 1], [0.4, 0.6]),
        #Consultant
        "consultant_prior_collaboration" : ([0, 1], [0.6, 0.4])
    }

    project_types = ["Industrieel", "Wonen", "Institutioneel", "Agrarisch",
                    "Civiel (infra)", "Gebied", "Commercieel"]
    project_type_probs = [0.08, 0.45, 0.10, 0.03, 0.15, 0.02, 0.17]
    project_type = np.random.choice(project_types, size=n_samples, p=project_type_probs)

    type_size_params = {
        "Industrieel":    {"mean": 40000, "std": 10000},
        "Wonen":          {"mean": 12000, "std": 5000},
        "Institutioneel": {"mean": 20000, "std": 7000},
        "Agrarisch":      {"mean": 8000,  "std": 3000},
        "Civiel (infra)": {"mean": 30000, "std": 8000},
        "Gebied":         {"mean": 10000, "std": 4000},
        "Commercieel":    {"mean": 18000, "std": 6000}
    }

    project_size = np.array([
        np.random.normal(type_size_params[t]["mean"], type_size_params[t]["std"])
        for t in project_type
    ])
    project_size = np.clip(project_size, 100, 100000)

    contract_value = (
        0.0012 * project_size + np.random.normal(0, 3, n_samples)
    )
    contract_value = np.clip(contract_value, 0.2, 300)

    project_days = (150 + 0.02 * project_size + np.random.normal(0, 100, n_samples))
    project_days = np.clip(project_days, 60, 2400)

    df_numerical = pd.DataFrame({
        "project_size_m2": project_size,
        "contract_value_million": contract_value,
        "planned_duration_days": np.round(project_days).astype(int)
    })
    for var, params in numerical_vars.items():
        data = np.random.normal(loc=params["mean"], scale=params["std"], size=n_samples)
        if var in ["planned_duration_days", "num_subcontractors"]:
            rounded = np.round(data).astype(int)
            df_numerical[var] = np.clip(rounded, params["min"], params["max"])
        else:
            df_numerical[var] = np.clip(data, params["min"], params["max"])

    categorical_vars["project_type"] = (project_types, project_type_probs)
    df_categorical = pd.DataFrame({"project_type": project_type})
    for var, (choices, probs) in categorical_vars.items():
        if var != "project_type":
            df_categorical[var] = np.random.choice(choices, size=n_samples, p=probs)
            
    df_num_norm = df_numerical.copy()
    df_cat_norm = df_categorical.copy()

    for var in df_num_norm.columns:
        df_num_norm[var] = (df_num_norm[var] - df_num_norm[var].min()) / (df_num_norm[var].max() - df_num_norm[var].min())

    award_method_map = {"Laagste prijs": 0.5, "Laagste kosten op basis van kosteneffectiviteit": 0.25, "Beste Prijs-Kwaliteitverhouding": 0}
    contract_type_map = {"Lump-sum": 0.5, "UAV-GC": 0.25, "Kost+": 0}

    df_cat_norm["contract_award_method"] = df_cat_norm["contract_award_method"].map(award_method_map)
    df_cat_norm["contract_type"] = df_cat_norm["contract_type"].map(contract_type_map)
    df_cat_norm["consultant_prior_collaboration"] = df_cat_norm["consultant_prior_collaboration"].astype(float)
    df_cat_norm["use_of_bim"] = df_cat_norm["use_of_bim"].astype(float)
    df_cat_norm["consultant_prior_collaboration"] = 0.125 + 0.25 * df_cat_norm["consultant_prior_collaboration"]
    df_cat_norm["use_of_bim"] = 0.125 + 0.25 * df_cat_norm["use_of_bim"]
    df_cat_norm = df_cat_norm.drop(columns=["project_type"])

    df_norm = pd.concat([df_num_norm, df_cat_norm], axis=1)

    group_vars = {
        "owner": [
            "payment_reliability",
            "decision_making_structure",
            "change_order_freq"
        ],
        "contractor": [
            "contractor_experience",
            "num_subcontractors",
            "planning_detail"
        ],
        "design": [
            "design_maturity",
            "pre_design_data_completeness",
            "design_team_experience_years",
            "use_of_bim"
        ],
        "labor": [
            "labour_shortage",
            "experienced_labour"
        ],
        "consultant": [
            "consultant_experience",
            "consultant_availability", 
            "consultant_prior_collaboration"
        ],
        "materials": [
            "material_supply_reliability",
            "expected_inflation"
        ],
        "external": [
            "weather_risk",
            "subsurface_risk",
            "permitting_complexity"
        ],
        "equipment": [
            "equipment_availability",
            "average_equipment_age"
        ],
        "project": [
            "project_size_m2",
            "contract_value_million",
            "contract_award_method",
            "contract_type"
        ]
    }

    var_direction = {
        "payment_reliability": -1,
        "decision_making_structure": +1,
        "change_order_freq": +1,
        "contractor_experience": -1,
        "num_subcontractors": +1,
        "planning_detail": -1,
        "design_maturity": -1,
        "pre_design_data_completeness": -1,
        "design_team_experience_years": -1,
        "use_of_bim": -1,
        "labour_shortage": +1,
        "experienced_labour": -1,
        "consultant_experience": -1,
        "consultant_availability": -1,
        "consultant_prior_collaboration": -1,
        "material_supply_reliability": -1,
        "expected_inflation": +1,
        "weather_risk": +1,
        "subsurface_risk": +1,
        "permitting_complexity": +1,
        "equipment_availability": -1,
        "average_equipment_age": +1,
        "project_size_m2": +1,
        "contract_value_million": +1,
        "contract_award_method": +1,
        "contract_type": +1
    }

    df_group_scores = pd.DataFrame()
    for group, vars_list in group_vars.items():
        weighted_vars = [(df_norm[v] * var_direction.get(v, 1)) for v in vars_list]
        df_group_scores[f"{group}_score"] = pd.concat(weighted_vars, axis=1).mean(axis=1)

    base_delay_prob = 0.53
    base_delay_severity = 0.55

    freq_index = {
        "owner": 61.39, "contractor": 56.81, "design": 56.45, "labor": 56.51,
        "consultant": 56.53, "materials": 49.55, "external": 48.9,
        "project": 48.96, "equipment": 48.43
    }

    sev_index = {
        "owner": 71.47, "contractor": 67.08, "design": 66.65, "labor": 66.58,
        "consultant": 66.49, "materials": 62.34, "external": 56.60,
        "project": 56.15, "equipment": 56.72
    }
    freq_norm = {k: v / max(freq_index.values()) for k, v in freq_index.items()}
    sev_norm = {k: v / max(sev_index.values()) for k, v in sev_index.items()}

    delay_prob_score = sum(df_group_scores[f"{g}_score"] * freq_norm[g] for g in freq_norm)
    delay_prob_score_norm = (delay_prob_score - delay_prob_score.min()) / (delay_prob_score.max() - delay_prob_score.min())
    delay_prob = base_delay_prob + (delay_prob_score_norm - delay_prob_score_norm.mean()) * 5
    delay_prob = np.clip(delay_prob, 0.05, 0.95)

    df_delay_contrib = pd.DataFrame()
    df_delay_contrib["is_delayed"] = np.random.binomial(1, delay_prob, size=len(delay_prob))

    delay_sev_score = sum(df_group_scores[f"{g}_score"] * sev_norm[g] for g in sev_norm)
    delay_sev_score_norm = (delay_sev_score - delay_sev_score.min()) / (delay_sev_score.max() - delay_sev_score.min())
    delay_pct = base_delay_severity + (delay_sev_score_norm - delay_sev_score_norm.mean()) * 1
    noise = np.random.normal(0, 0.02, len(delay_pct))
    delay_pct = np.clip(delay_pct + noise, 0, 6)

    df_delay_contrib["delay_pct"] = df_delay_contrib["is_delayed"] * delay_pct

    mean_delayed = df_delay_contrib.loc[df_delay_contrib["is_delayed"] == 1, "delay_pct"].mean()
    df_delay_contrib.loc[df_delay_contrib["is_delayed"] == 1, "delay_pct"] *= base_delay_severity / mean_delayed

    df_target = pd.DataFrame({
        "is_delayed": df_delay_contrib["is_delayed"].astype(int),
        "delay_pct": df_delay_contrib["delay_pct"].astype(float)
    })
    df_target["delay_days"] = df_target["delay_pct"] * df_numerical["planned_duration_days"]

    df = pd.concat([df_numerical.reset_index(drop=True), df_categorical.reset_index(drop=True), df_target.reset_index(drop=True)], axis=1)

    st.write("Data gegenereerd. Weergave van de eerste vijf gegenereerde projecten:")
    st.dataframe(df.head().rename(columns=lambda c: variable_labels.get(c, c)))

    csv = df.to_csv(index=False).encode('utf-8')
    st.download_button(
        label="Download CSV",
        data=csv,
        file_name='synthetic_project_data.csv',
        mime='text/csv'
    )

    return df, df_numerical, df_categorical, categorical_vars

def generate_template():
    numerical_cols = [
        "payment_reliability", "decision_making_structure", "change_order_freq",
        "contractor_experience", "num_subcontractors", "planning_detail",
        "design_maturity", "pre_design_data_completeness", "design_team_experience_years",
        "labour_shortage", "experienced_labour", "consultant_experience",
        "consultant_availability", "material_supply_reliability", "expected_inflation",
        "weather_risk", "subsurface_risk", "permitting_complexity",
        "equipment_availability", "average_equipment_age",
        "project_size_m2", "contract_value_million", "planned_duration_days", "delay_pct"
    ]

    categorical_cols = [
        "project_type", "contract_award_method", "contract_type",
        "use_of_bim", "consultant_prior_collaboration", "is_delayed"
    ]

    categorical_options = {
        "project_type": ["Industrial", "Residential", "Institutional", "Agricultural", "Civil", "Environmental", "Commercial"],
        "contract_award_method": ["Lowest bid", "Negotiated", "Quality-based"],
        "contract_type": ["Lump-sum", "Design-build", "Cost-plus"],
        "use_of_bim": [0, 1],
        "consultant_prior_collaboration": [0, 1],
        "is_delayed": [0, 1]
    }

    all_cols = numerical_cols + categorical_cols

    df_template = pd.DataFrame(columns=all_cols)
    for col in numerical_cols:
        df_template[col] = pd.Series(dtype=float)
    for col in categorical_cols:
        df_template[col] = pd.Series(dtype=str)

    df_template.loc[0] = [
        3, 5, 10, 15, 5, 85, 70, 80, 10, 8, 60, 10, 90, 4, 3,
        6, 3, 3, 0.9, 5, 20000, 20, 365, 0.1, "Residential", "Lowest bid",
        "Lump-sum", "1", "1", "1"
    ]

    return df_template, categorical_options

def generate_template_excel():
    df_template, categorical_options = generate_template()  

    file_name = "project_data_template.xlsx"
    df_template.to_excel(file_name, index=False)

    wb = load_workbook(file_name)
    ws = wb.active

    for col_idx, col_name in enumerate(df_template.columns, start=1):
        if col_name in categorical_options:
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(map(str, categorical_options[col_name]))}"',
                allow_blank=True
            )
            ws.add_data_validation(dv)
            dv.add(f"{ws.cell(row=2, column=col_idx).coordinate}:{ws.cell(row=100, column=col_idx).coordinate}")

    for col_idx, col_name in enumerate(df_template.columns, start=1):
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0 
            for cell in ws[ws.cell(row=1, column=col_idx).column_letter]
        )
        max_length = max(max_length, len(col_name)) + 2  # add some padding
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length

    wb.save(file_name)
    return file_name